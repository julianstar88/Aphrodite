# -*- coding: utf-8 -*-
"""
Created on Tue Mar  3 22:30:14 2020

@author: Julian
"""
import datetime
import pathlib2
import openpyxl
import numpy as np
import tempfile
import comtypes.client as com
from PyQt5 import QtGui

import MainModules.Database as db
import UtilityModules.ExporterUtilityModules as exporterUtils
from UtilityModules.CustomModel import CustomSqlModel
from UtilityModules.MiscUtilities import ModelInputValidation


class Exporter():
    """
    exports a table from a database to an excel-file (.xlsx). although the
    database can be set arbitrarily, the database is supposed to be a
    training-routine-database.
    this class is meant to be used in the framwork of 'Aphrodite'

    Properties
    ----------
    - database:
        a path to a existing database file, which is a valid trainingroutine

        - default: None
        - getter: database()
        - setter: setDatabase(db-file)

    - exportPath:
        a path to a dirctory to store the created exportfile in

        - default: None
        - getter: exportPath()
        - setter: setExportPath(directory-path)

    - model:
        a refenrence to a valid CustomSqlModel

        - default: None
        - getter: model()
        - setter: setModel(model)

    - name:
        username for the trainingroutine

        - default: None
        - getter: name()
        - setter: setName(name)

    - routineName:
        name of the exportfile

        - default: None
        - getter: routineName()
        - setter: setRoutineName(routine-name)

    - trainingPeriode:
        discribes the duration for the trainingroutine. One has only to provide
        the start date via the setter method, the end date will be calulated
        automatically six weeks later. the getter will return a list consisting
        of strings describing the start date in first place and the end date in
        last place

        - default: None
        - getter: trainingPeriode()
        - setter: setTrainingPeriode(year, month, day)

    - trainingMode:
        describes which kind of trainingform is written into the exportfile

        - default: None
        - getter: trainingMode
        - setter: setTrainingMode(mode)

    - workBook:
        'workBook' is an openpyxl.Workbook object, used to create the exportfile.
        this property gets written, if one set a path to a valid database. it
        also can be changed using its setter method

        - default: none
        - getter: workBook()
        - setter: setWorkBook(workbook)

    Usage
    -----

    1. create the exporter object:

        >>> exporter = Exporter()

    2. set suitable properties:

        >>> exporter.setDatabase(...)
        >>> exporter.setExportPath(...)
        >>> exporter.setName(...)
        >>> exporter.setRoutineName(...)
        >>> exporter.setTrainingPeriode(...)
        >>> exporter.setTrainingMode(...)

        >>> exporter.setRoutineModel(...)
        >>> exporter.setAlternativeModel(...)
        >>> exporter.setNoteModel(...)


    3. layout the exportfile:

        >>> exporter.routineLayout()

    4. populate the exportfile with exercises, sets and repetitions

        either with data from Database:

        >>> exporter.populateRoutine(exporter.dataFromDatabase(...))

        or, if routineModel, alternativeModel and noteModel have been set

        >>> exporter.populateRoutine(exporter.dataFromModel(...))

    5. save the workBook to create the exportfile

        >>> exporter.saveRoutine()

    """

    def __init__(self,
                 database = None,
                 exportPath = None,
                 name = None,
                 routineName = None,
                 startDate = None,
                 trainingMode = None,
                 workBook = None,
                 routineModel = None,
                 alternativeModel = None,
                 noteModel = None):

        self._database = None
        self._databaseName = None
        self._databasePath = None
        self._exportPath = None
        self._name = None
        self._routineName = None
        self._trainingPeriode = list()
        self._trainingMode = None
        self._workBook = None
        self._routineModel = None
        self._alternativeModel = None
        self._noteModel = None

        # purly private properties without getter or setter.
        # this properties are uesed to finalize the layout (set superscripts
        # and subscripts) after saving the epxort file.
        self.__routineStartRow = int()
        self.__alternativeStartRow = int()
        self.__noteStartRow = int()

        if (database):
            self.setDatabase(database)
        if (exportPath):
            self.setExportPath(exportPath)
        if (name):
            self.setName(name)
        if (routineName):
            self.setRoutineName(routineName)
        if (isinstance(startDate, tuple)) and (len(startDate) == 3):
            self.setTrainingPeriode(
                    startDate[0],
                    startDate[1],
                    startDate[2]
                )
        if (trainingMode):
            self.setTrainingMode(trainingMode)
        if (workBook):
            self.setWorkBook(workBook)
        if (routineModel):
            self.setRoutineModel(routineModel)
        if (alternativeModel):
            self.setAlternativeModel(alternativeModel)
        if (noteModel):
            self.setNoteModel(noteModel)

    def alternativeModel(self):
        """
        getter method for the attribute 'alternativeModel'

        Returns
        -------
        CustomSqlModel or QtGui.QStandardItemModel

        """
        return self._alternativeModel

    def database(self):
        """
        holds thepath to a existing database file, which is a valid
        trainingroutine. the database is needed to populate the the
        exportfile.


        Returns
        -------
        str
            path to a database.

        """

        return self._database

    def databaseName(self):
        """
        holds the name of the database. this property is written by
        'setDatabase'. it is the equivalent of invoking
        'pathlib2.Path(database).stem'

        Returns
        -------
        str
            databaseName.

        """

        return self._databaseName

    def databasePath(self):
        """
        holds the directory of 'databaseName'. this property is written by
        'setDatabase'. it is the equivalent of invoking str(pathlib2.Path(database).parent)

        Returns
        -------
        str
            databasePath.

        """

        return self._databasePath

    def dataFromDatabase(self, database = None):
        """
        retrieve data from a database as source for training data. the property
        'database' can be st either by calling 'setDatabase' or directly by
        calling 'dataFromDatabase(database)' with a path to a valid database-file.
        additionally, one can set the tabel from wich to retrive data, by setting
        the 'tableName' argument to a valid table.

        Parameters
        ----------
        database : str, optional
            path to a valid database-file. The default is None.

        Raises
        ------
        TypeError
            will be raised, if no valid value for the database-property
            has been set.

        Returns
        -------
        routineData : list
            all data from training_routine  as nested list.

        alternativeData : list
            all data from training_alternatives as nested list.

        informationData : list
            all data from general_information as nested list.

        """

        if database:
            self.setDatabase(database)

        if not self.database():
            raise TypeError(
                    "before fetching data from a database, set a path to a valid database-file"
                )
        databaseObj = db.database(self.database())
        routineData = databaseObj.data("training_routine")
        alternativeData = databaseObj.data("training_alternatives")
        noteData = databaseObj.data("training_notes")

        return routineData, alternativeData, noteData

    def dataFromModel(
                self,
                routineModel = None,
                alternativeModel = None,
                noteModel = None):
        """
        retrieve data from different Models as source for routine-data,
        alternative-data and note-data and returns the data as lists.

        Parameters
        ----------
        routineModel : CustomSqlModel or QtGui.QStandardItemModel, optional
            data source for routine data. The default is None.

        alternativeModel : CustomSqlModel or QtGui.QStandardItemModel, optional
            data source for alternative data. The default is None.

        noteModel : CustomSqlModel or QtGui.QStandardItemModel, optional
            data source for model data. The default is None.

        Raises
        ------
        TypeError
            will be raised, if the values for the properties does not match
            <CustomSqlModel> or <QtGui.QStandardItemModel>.
            EXCEPTION: noteModel must explicitly be <QtGui.QStandardItemModel>

        Returns
        -------
        routineData : list
            all data from the routineModel as nested list.

        alternativeData : list
            all data from the alternativeModel as nested list.

        noteData : list
            all data from the noteModel as nested list.

        """

        if routineModel:
            self.setRoutineModel(routineModel)
        if alternativeModel:
            self.setAlternativeModel(alternativeModel)
        if noteModel:
            self.setNoteModel(noteModel)

        if (not isinstance(self.routineModel(), CustomSqlModel)) and \
            (not isinstance(self.routineModel(), QtGui.QStandarItemModel)):
            raise TypeError(
                    "before fetching data from a model, a valid model has to be set for 'routineModel'"
                )
        if (not isinstance(self.alternativeModel(), CustomSqlModel)) and \
            (not isinstance(self.alternative(), QtGui.QStandarItemModel)):
            raise TypeError(
                    "before fetching data from a model, a valid model has to be set for 'alternativeModel'"
                )
        if (not isinstance(self.noteModel(), QtGui.QStandardItemModel)):
            raise TypeError(
                    "before fetching data from a model, a valid model has to be set for 'noteModel'"
                )

        # retrieve routine Data
        rows = self.routineModel().rowCount()
        cols = self.routineModel().columnCount()
        routineData = list()
        for row in range(rows):
            line = []
            for col in range(cols):
                item = self.routineModel().item(row, col)
                line.append(item.userData())
            routineData.append(line)

        # retrieve alternative Data
        rows = self.alternativeModel().rowCount()
        cols = self.alternativeModel().columnCount()
        alternativeData = list()
        for row in range(rows):
            line = []
            for col in range(cols):
                item = self.alternativeModel().item(row, col)
                line.append(item.userData())
            alternativeData.append(line)

        # retrieve note Data
        rows = self.noteModel().rowCount()
        cols = self.noteModel().columnCount()
        noteData = list()
        for row in range(rows):
            line = []
            for col in range(cols):
                item = self.noteModel().item(row, col)
                line.append(item.userData())
            noteData.append(line)
        return routineData, alternativeData, noteData

    def exportPath(self):
        """
        holds the directory-path where to store the exportfile

        Returns
        -------
        str
            exportPath.

        """

        return self._exportPath

    def finalizeLayout(self, file):

        routineData, alternativeData, noteData = self.dataFromDatabase()
        file = pathlib2.Path(file)
        app = com.CreateObject("Excel.Application")

        # for debugging purposes can this property set to True
        app.Visible = False

        wb = app.Workbooks.Open(str(file))
        ws = wb.worksheets[1]

        for i in range(len(routineData)):
            rowID = i + 1

            l = list()
            for alternative in alternativeData:
                if rowID == alternative[0]:
                    l.append(alternative[1])
            alternatives = "%s"*len(l)
            alternatives = alternatives % tuple(l)

            l = list()
            for note in noteData:
                if rowID == note[0]:
                    l.append(note[1])
            notes = "%s"*len(l)
            notes = notes % tuple(l)

            base = routineData[i][0]

            newValue = "{baseName}{alternatives}{notes}".format(
                    baseName = base,
                    alternatives = alternatives,
                    notes = notes
                )

            cell = ws.Cells[self.__routineStartRow + i, 1]
            cell.Value[:] = newValue

            # set superscript
            pos = len(base) + 1
            length = len(alternatives)
            if length > 0:
                cell.Characters[pos, length].Font.Superscript = True

            # set subscript
            pos = len(base) + len(alternatives) + 1
            length = len(notes)
            if length > 0:
                cell.Characters[pos, length].Font.Subscript = True

        wb.Save()
        app.Quit()
        return True


    def name(self):
        """
        holds the username for the trainingroutine. this is the name, which
        will be written under the name in the exportfile

        Returns
        -------
        str
            name.

        """

        return self._name

    def noteModel(self):
        """
        getter method for the attribute 'noteModel'

        Returns
        -------
        QtGui.QStandardItemModel

        """
        return self._noteModel

    def populateRoutine(self, routineData, alternativeData, noteData):
        """
        by invoking this method, the workbook in the 'workBook' property gets
        populated with data from routineData, alternativeData and noteData.
        These arguments are typically set by 'dataFromDatabase' or
        'dataFromModel'

        Note:
        make sure, that 'routineLayout' has been invoked prior.

        Note:
        if no proper values either for the database-property or the
        workBook-property have been set, a ValueError will be raised

        Parameters
        ----------
        routineData: list
            a list with values representing the trainingroutine

        alternativeData: list
            a list with values representing the trainingalternatives

        noteData: list
            a list with values representing the trainingnotes

        Raises
        ------
        TypeError
             will be raised if the input is not type list or type tuple, or the
             'workBook' property does not hold a valid 'openpyxl.Workbook' object

        ValueError
            will be raised, if the dimension of the 'numpy.array' representation
            of the input arguments does not match (n, m)

        Returns
        -------
        None.

        """
        if not isinstance(routineData, tuple) and not isinstance(routineData, list):
            raise TypeError(
                    "input <{input_type}> does not match {expected_type_1} or {expected_type_2} for argument 'routineData'".format(
                            input_type = type(routineData),
                            expected_type_1 = tuple,
                            expected_type_2 = list
                        )
                )

        if not isinstance(alternativeData, tuple) and not isinstance(alternativeData, list):
            raise TypeError(
                    "input <{input_type}> does not match {expected_type_1} or {expected_type_2} for argument 'alternativeData'".format(
                            input_type = type(alternativeData),
                            expected_type_1 = tuple,
                            expected_type_2 = list
                        )
                )

        if not isinstance(noteData, tuple) and not isinstance(noteData, list):
            raise TypeError(
                    "input <{input_type}> does not match {expected_type_1} or {expected_type_2} for argument 'noteData'".format(
                            input_type = type(noteData),
                            expected_type_1 = tuple,
                            expected_type_2 = list
                        )
                )

        if not isinstance(self.workBook(), openpyxl.Workbook):
            raise TypeError(
                    "tried to access an invalid workbook. set a valid openpyxl.Workbook-object as workbook, before populating a trainingroutine "
                )

        ws = self.workBook().active
        validator = ModelInputValidation()

        # set header data
        ws["A3"] = self.name()
        ws["F3"] = self.trainingPeriode()[0]
        ws["F4"] = self.trainingPeriode()[1]
        ws["H3"] = self.trainingMode()

        self.__routineStartRow = 7
        exporterUtils.setAlignment(
                ws,
                exporterUtils.generateRangeExpression(
                        startRow = self.__routineStartRow,
                        endRow = len(list(ws.rows)),
                        startColumn = "C",
                        endColumn = "J"
                    ),
                horizontal = "left"
            )

        """set routine data"""
        inputValues = [routineData[i][0] for i in range(len(routineData))]

        # exercies names
        for i, val in enumerate(inputValues):
            rowID = i + 1
            ws["A" + str(self.__routineStartRow + i)] = val
            l = list()
            for alternative in alternativeData:
                if rowID == alternative[0]:
                    l.append(alternative[1])
            alternatives = "%s)"*len(l)
            alternatives = alternatives % tuple(l)

            l = list()
            for note in noteData:
                if rowID == note[0]:
                    l.append(note[1])
            notes = "%s)"*len(l)
            notes = notes % tuple(l)

            ws["A" + str(self.__routineStartRow + i)] = val + alternatives + notes

        # exercise values
        for n, row in enumerate(routineData):
            row = row[1:-1]
            del row[2]
            for m, val in enumerate(row):

                # if values are readable as numeric values in a
                # 'ModelInputValidation' manner, convert them into integer
                # (prevent excel from throwing a warning for writing numbers
                # as text)
                if validator.checkValue(val):
                    val = validator.readValue(val)[0]

                ws.cell(
                        row = self.__routineStartRow + n,
                        column = 3 + m,
                        value = val
                    )

        """set alternative data"""
        self.__alternativeStartRow = self.__routineStartRow + len(routineData) + 3
        ws.cell(self.__alternativeStartRow, 1, value = "Alternativen:").font = openpyxl.styles.Font(
            b = True
        )
        self.__alternativeStartRow += 1
        inputValues = [alternativeData[i][1:4] for i in range(len(alternativeData))]

        # alternative exercise names
        for i, val in enumerate(inputValues):
            ws["A" + str(self.__alternativeStartRow + i)] = val[0] + ") " + val[2]

        # alternative exercise values
        for n, row in enumerate(alternativeData):
            row = row[4:-1]
            del row[2]
            for m, val in enumerate(row):

                # if values are readable as numeric values in a
                # 'ModelInputValidation' manner, convert them into integer
                # (prevent excel from throwing a warning for writing numbers
                # as text)
                if validator.checkValue(val):
                    val = validator.readValue(val)[0]

                ws.cell(
                        row = self.__alternativeStartRow + n,
                        column = 3 + m,
                        value = val
                    )

        """set note data"""
        self.__noteStartRow = len(list(ws.rows)) + 1
        ws.cell(self.__noteStartRow, 1, value = "Notes:").font = openpyxl.styles.Font(
            b = True
        )
        self.__noteStartRow += 1

        # note name
        inputValues = [noteData[i][1] for i in range(len(noteData))]
        for i, val in enumerate(inputValues):
            ws["A" + str(self.__noteStartRow + i)] = val + ")"

        # note value
        inputValues = [noteData[i][3] for i in range(len(noteData))]
        for i, val in enumerate(inputValues):
            ws.merge_cells(exporterUtils.generateRangeExpression(
                    self.__noteStartRow + i,
                    self.__noteStartRow + i,
                    "B",
                    "J"
                ))
            ws["B" + str(self.__noteStartRow + i)] = val

    def routineLayout(self, rows = 40):
        """
        by invoking this method, the layout for the trainingroutine
        in the exportfile will be created. its also writing the intermediate
        workbook file to the 'workBook'-property. for convinience, the created
        intermediade workbook will be returned to the user.

        Parameters
        ----------
        rows : int, optional
            determines, how many rows in the exportfile will be available.
            The default is 40.

        Returns
        -------
        workBook : openpyxl.Workbook
            intermediate workbook ojbect.

        """

        workBook = exporterUtils.TamplateLayout(rows)
        self.setWorkBook(workBook)
        return workBook

    def routineModel(self):
        """
        holds a reference to a vild CustomSqlModel representing routine-data
        of a trainingroutine

        Returns
        -------
        CustomSqlModel

        """
        return self._routineModel

    def routineName(self):
        """
        holds the routinename for the exportfile. this will be the name for the
        -excel-file, which will be saved at 'exportPath'

        Returns
        -------
        str
            routineName.

        """

        return self._routineName

    def trainingMode(self):
        """
        holds the mode, which discribes the kind of training. this mode will be
        written into the exportfile

        Returns
        -------
        str
            trainingMode.

        """

        return self._trainingMode

    def trainingPeriode(self):
        """
        holds the duration for the trainingroutine. the dates in this property
        will be written to the exportfile

        Returns
        -------
        tuple
            contains the start and end date of the given trainingroutine.

        """

        return self._trainingPeriode

    def saveRoutine(self):
        """
        save the intermediate workbook 'workBook' and create the final export-
        file (.xlsx-file). the exportfile will be saved to the 'exportPath',
        with the name 'routineName'. to prevent unexpected behavior, one should
        obey the recommended procedure in the class-documentation.

        Raises
        ------
        TypeError
            will be raised, if either 'exportPath' or 'routineName' are
            invalid.

        Returns
        -------
        None.

        """


        if not self.exportPath():
            raise TypeError (
                    "try to export a trainingroutine to an invalid path. set a valid export path before saving a trainingroutine"
                )

        if not self.routineName():
            raise TypeError(
                    "try to export a trainingroutine with invalid routine name. set a valid routine name before saving a trainingroutine"
                )
        path = pathlib2.Path(self.exportPath()) / pathlib2.Path(self.routineName())
        self.workBook().save(str(path))

    def setAlternativeModel(self, model):
        """
        setter method for the attribute 'alternativeModel'

        Parameters
        ----------
        model : CustomSqlModel or QtGui.QStandardItemModel

        Raises
        ------
        TypeError
            will be raised, if <model> does not match <CustomSqlModel> or <QtGui.QStandardItemModel>.

        Returns
        -------
        None.

        """
        if (not isinstance(model, QtGui.QStandardItemModel)) and (not isinstance(model, CustomSqlModel)):
            raise TypeError(
                    "input <{input_name}> does not match {input_name_1} or {input_name_2}".format(
                            input_name = str(model),
                            type_name_1 = QtGui.QStandardItemModel,
                            type_name_2 = CustomSqlModel
                        )
                )

        self._alternativeModel = model

    def setDatabase(self, databasePath):
        """
        setter method for the 'database'-property.

        Parameters
        ----------
        databasePath : str; pathlib2.Path()
            this parameter must point to a valid db-file
            (typically a trainingroutine).

        Raises
        ------
        TypeError
             will be raised if the input is not type str.

        ValueError
            will be raised if input is an empty string or a path which
            does not point to a file.

        Returns
        -------
        None.

        """

        if not isinstance(databasePath, str) and not isinstance(databasePath, pathlib2.Path):
            raise TypeError(
                    "input <{input_value}> for argument 'databasePath' does not match {type_name}".format(
                            input_value = databasePath,
                            type_name = type("str")
                        )
                )
        path = pathlib2.Path(databasePath)
        if not path.exists():
            raise ValueError(
                    "Database does not exist: {input_name}".format(
                            input_name = str(path)
                        )
                )
        if str(path) == ".":
            raise ValueError(
                    "invalid input for argument 'databasePath'"
                )
        self._database = str(path)
        self._databaseName = path.stem
        self._databasePath = path.parent

    def setExportPath(self, exportPath):
        """
        setter method for the 'exportPath'-property.

        Parameters
        ----------
        exportPath : str, pathlib2.Path
            exportPath must point to a directory, where the exportfile
            will be stored.

        Raises
        ------
        TypeError
            will be raised, if the input is not type str.

        ValueError
            will be raised, if the input does not point to a directory,
            or the length of the input string is zeor.

        Returns
        -------
        None.

        """

        path = pathlib2.Path(exportPath)
        if not type(exportPath) == str:
            raise TypeError(
                    "input for argument 'exportPath' does not match {type_name}".format(
                            input_name = path,
                            type_name = type("str")
                        )
                )
        if not path.is_dir():
            raise ValueError(
                    "input for argument 'exportPath' does not refer to an existing directory"
                )
        if len(exportPath) == 0:
            raise ValueError(
                    "invalid input for argument 'exportPath'"
                )
        self._exportPath = str(path)

    def setModel(self, model):
        """
        setter method for the property: model

        Parameters
        ----------
        model : CustomSqlModel
            valid reference to a CustomSqlModel-object.

        Raises
        ------
        TypeError
            will be raised, if model is no CustomSqlModel-type.

        Returns
        -------
        None.

        """

        if not isinstance(model, CustomSqlModel):
            raise TypeError(
                    "input {input_name} for 'setModel' does not match {type_name}".format(
                            input_name = type(model),
                            type_name = CustomSqlModel
                        )
                )

        self._model = model

    def setName(self, name):
        """
        setter methof for the 'name'-property.

        Parameters
        ----------
        name : str
            Username, written in the exportfile.

        Raises
        ------
        TypeError
            will be raised, if the input is not type str.

        ValueError
            will be raised, if the length of the input is zero.

        Returns
        -------
        None.

        """

        if not isinstance(name, str):
            raise TypeError(
                "input <{name}> does not match {type_name}".format(
                        name = str(name),
                        type_name = str,
                    )
                )
        if len(name) == 0:
            raise ValueError("empty input for the attribute 'name'")

        self._name = name

    def setNoteModel(self, model):
        """
        setter method for the attribute 'noteModel'

        Parameters
        ----------
        model : QtGui.QStandardItemModel

        Raises
        ------
        TypeError
            will be raised if <model> is not type <QtGui.QStandardItemModel>.

        Returns
        -------
        None.

        """
        if (not isinstance(model, QtGui.QStandardItemModel)) and (not isinstance(model, CustomSqlModel)):
            raise TypeError(
                    "input <{input_name}> does not match {input_name_1} or {input_name_2}".format(
                            input_name = str(model),
                            type_name_1 = QtGui.QStandardItemModel,
                            type_name_2 = CustomSqlModel
                        )
                )

        self._noteModel = model

    def setRoutineModel(self, model):
        """
        set the attribute 'routineModel' to <model>

        Parameters
        ----------
        model : CustomSqlModle, QtGui.QStandartItemModel

        Raises
        ------
        TypeError
            will be raised if <model> is not type 'CustomSqlModel' or 'QtGui.QStandardItemModel'

        Returns
        -------
        None.

        """
        if (not isinstance(model, QtGui.QStandardItemModel)) and (not isinstance(model, CustomSqlModel)):
            raise TypeError(
                    "input <{input_name}> does not match {type_name}".format(
                            input_name = str(model),
                            type_name = QtGui.QStandardItemModel
                        )
                )

        self._routineModel = model

    def setRoutineName(self, routineName):
        """
        setter method for the 'routineName'-property

        Parameters
        ----------
        routineName : str
            this represents the name of the exporfile.

        Raises
        ------
        TypeError
            will be raised.
        ValueError
            DESCRIPTION.

        Returns
        -------
        None.

        """

        if not type(routineName) == str:
            raise TypeError(
                "input argument {name} does not match {type_name}".format(
                        name = routineName,
                        type_name = type("str")
                    )
                )
        elif len(routineName) == 0:
            raise ValueError("empty input for the attribute 'routineName'")
        else:
            self._routineName = routineName

    def setTrainingMode(self, trainingMode):
        """
        setter for the 'trainingMode'-property

        Parameters
        ----------
        trainingMode : str
            descirbes which the type of training. this property will be written
            inot the exportfile.

        Raises
        ------
        TypeError
            will be raised, if the input is not type str.
        ValueError
            will be raised, if input length is zero.

        Returns
        -------
        None.

        """

        if not type(trainingMode) == str:
            raise TypeError(
                    "input {input_name} for argument 'trainingMode' does not match {type_name}".format(
                            input_name = type(trainingMode),
                            type_name = type("123")
                        )
                )
        if len(trainingMode) == 0:
            raise ValueError(
                    "input for 'trainingMode' is not valid. 'trainingMode' must at least contain one character"
                )
        self._trainingMode = trainingMode

    def setTrainingPeriode(self, startYear, startMonth, startDay,
                           endYear = None, endMonth = None, endDay = None):
        """
        setter method for the 'trainingPeriode'-property. the setter provides
        only the possibility to set the start date. the end date will be
        computed automatically six weeks later.

        Parameters
        ----------
        startYear : int
            start year of the training.
        startMonth : int
            start month of the training.
        startDay : int
            start day of the training.

        Raises
        ------
        TypeError
            will be raised, if one of the parameters is not type int.

        Returns
        -------
        None.

        """

        if not type(startYear) == int:
            raise TypeError(
                    """input argument '{name}' for <startYear>
                    does not match {type_name}""".format(
                            name = startYear,
                            type_name = type(123)
                        )
                )
        if not type(startMonth) == int:
            raise TypeError(
                    """input argument '{name}' for <startYear>
                    does not match {type_name}""".format(
                            name = startMonth,
                            type_name = type(123)
                        )
                )
        if not type(startDay) == int:
            raise TypeError(
                    """input argument '{name}' for <startYear>
                    does not match {type_name}""".format(
                            name = startDay,
                            type_name = type(123)
                        )
                )
        startDate = datetime.date(startYear, startMonth, startDay)

        if endYear and endMonth and endDay:
            endDate = datetime.date(endYear, endMonth, endDay)
        else:
            endDate = startDate + datetime.timedelta(days = 42)
        startDateString = startDate.strftime("%d.%m.%Y")
        endDateString = endDate.strftime("%d.%m.%Y")
        self._trainingPeriode = [startDateString, endDateString]

    def setWorkBook(self, workBook):
        """
        setter method for the 'workBook'-property. this property is normally
        not meant to be set manually, because it will be set by 'setDatabase'.
        However, for convinience of special cases it is possible to modify the
        value of the 'workBook' later on

        Parameters
        ----------
        workBook : openpyxl.Workbook
            workbook objects will used as intermediate to create the exportfile.

        Raises
        ------
        TypeError
            will be raised, if the paramter is not type openpyxl.Workbook.

        Returns
        -------
        None.

        """

        if not type(workBook) == openpyxl.Workbook:
            raise TypeError(
                    "input for 'workBook' must be an instance of the 'openpyxl.Workbook' module"
                )
        self._workBook = workBook

    def workBook(self):
        """
        holds the intermediate workbook-object which will be needed to create
        the export file

        Returns
        -------
        openpyxl.Workbook
            the 'workBook'-property.

        """

        return self._workBook

if __name__ == "__main__":
    pass