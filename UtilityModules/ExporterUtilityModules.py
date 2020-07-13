# -*- coding: utf-8 -*-
"""
Created on Mon Apr 13 20:54:40 2020

@author: Julian
"""
import openpyxl
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from Utility_Function_Library.converter import ColorConverter

def setAlignment(ws, cellRange,
                 horizontal = "center",
                 vertical = "center"):

    cellRange = ws[cellRange]
    for row in cellRange:
        for cell in row:
            cell.alignment = Alignment(
                    horizontal = horizontal,
                    vertical = vertical
                )

def setBackground(ws, cellRange,
                  fill_type = None,
                  fgColor = Color(),
                  bgColor = Color(),
                  start_color = None,
                  end_color = None):

    cellRange = ws[cellRange]
    for i, row in enumerate(cellRange):
        for n, cell in enumerate(row):
            cell.fill = PatternFill(
                    fill_type = fill_type,
                    fgColor = fgColor,
                    bgColor = bgColor,
                    start_color = start_color,
                    end_color = end_color
                )

def setBorder(ws, cellRange,
              borderStyle = "thick",
              gridStyle = None):

    cellRange = ws[cellRange]

    if len(cellRange) == 1:
        row = cellRange[0]
        for i, cell in enumerate(row):
            if i == 0:
                cell.border = Border(
                        top = Side(
                                border_style = borderStyle
                            ),
                        left = Side(
                                border_style = borderStyle
                            ),
                        bottom = Side(
                                border_style = borderStyle
                            )
                    )
                if gridStyle:
                    cell.border = Border(
                            top = Side(
                                    border_style = borderStyle
                                ),
                            left = Side(
                                    border_style = borderStyle
                                ),
                            bottom = Side(
                                    border_style = borderStyle
                                ),
                            right = Side(
                                    border_style = gridStyle
                                )
                        )
            elif i == len(row)-1:
                cell.border = Border(
                        top = Side(
                                border_style = borderStyle
                            ),
                        bottom = Side(
                                border_style = borderStyle
                            ),
                        right = Side(
                                border_style = borderStyle
                            )
                    )
            else:
                cell.border = Border(
                        top = Side(
                                border_style = borderStyle
                            ),
                        bottom = Side(
                                border_style = borderStyle
                            )
                    )
                if gridStyle:
                    cell.border = Border(
                            top = Side(
                                    border_style = borderStyle
                                ),
                            bottom = Side(
                                    border_style = borderStyle
                                ),
                            right = Side(
                                    border_style = gridStyle
                                )
                        )
        return

    for i, row in enumerate(cellRange):
        for n, cell in enumerate(row):
            if i == 0:
                if n == 0:
                    cell.border = Border(
                            top = Side(
                                    border_style = borderStyle
                                ),
                            left = Side(
                                    border_style = borderStyle
                                )
                        )
                    if gridStyle:
                        cell.border = Border(
                                top = Side(
                                        border_style = borderStyle
                                    ),
                                left = Side(
                                        border_style = borderStyle
                                    ),
                                bottom = Side(
                                        border_style = gridStyle
                                    ),
                                right = Side(
                                        border_style = gridStyle
                                    )
                            )
                elif n == len(row)-1:
                    cell.border = Border(
                            top = Side(
                                    border_style = borderStyle
                                ),
                            right = Side(
                                    border_style = borderStyle
                                )
                        )
                    if gridStyle:
                        cell.border = Border(
                                top = Side(
                                        border_style = borderStyle
                                    ),
                                right = Side(
                                        border_style = borderStyle
                                    ),
                                bottom = Side(
                                        border_style = gridStyle
                                    )
                            )
                else:
                    cell.border = Border(
                            top = Side(
                                    border_style = borderStyle
                                )
                        )
                    if gridStyle:
                        cell.border = Border(
                                top = Side(
                                        border_style = borderStyle
                                    ),
                                bottom = Side(
                                        border_style = gridStyle
                                    ),
                                right = Side(
                                        border_style = gridStyle
                                    )
                            )
            elif i == len(cellRange)-1:
                if n == 0:
                    cell.border = Border(
                        bottom = Side(
                                border_style = borderStyle
                            ),
                        left = Side(
                                border_style = borderStyle
                            )
                        )
                    if gridStyle:
                        cell.border = Border(
                            bottom = Side(
                                    border_style = borderStyle
                                ),
                            left = Side(
                                    border_style = borderStyle
                                ),
                            right = Side(
                                    border_style = gridStyle
                                )
                            )
                elif n == len(row)-1:
                    cell.border = Border(
                            bottom = Side(
                                    border_style = borderStyle
                                ),
                            right = Side(
                                    border_style = borderStyle
                                )
                        )
                else:
                    cell.border = Border(
                            bottom = Side(
                                    border_style = borderStyle
                                )
                        )
                    if gridStyle:
                        cell.border = Border(
                                bottom = Side(
                                        border_style = borderStyle
                                    ),
                                right = Side(
                                        border_style = gridStyle
                                    )
                            )
            else:
                if n == 0:
                    cell.border = Border(
                            left = Side(
                                    border_style = borderStyle
                                )
                        )
                    if gridStyle:
                        cell.border = Border(
                                left = Side(
                                        border_style = borderStyle
                                    ),
                                bottom = Side(
                                        border_style = gridStyle
                                    ),
                                right = Side(
                                        border_style = gridStyle
                                    )
                            )
                elif n == len(row)-1:
                    cell.border = Border(
                            right = Side(
                                    border_style = borderStyle
                                )
                        )
                    if gridStyle:
                        cell.border = Border(
                                right = Side(
                                        border_style = borderStyle
                                    ),
                                bottom = Side(
                                        border_style = gridStyle
                                    )
                            )
                else:
                    if gridStyle:
                        cell.border = Border(
                                bottom = Side(
                                        border_style = gridStyle
                                    ),
                                right = Side(
                                        border_style = gridStyle
                                    )
                            )

def setGrid(ws, cellRange,
                  gridStyle = "thin"):

    cellRange = ws[cellRange]
    for i, row in enumerate(cellRange):
        for n, cell in enumerate(row):
            if i == len(cellRange)-1:
                if not n == len(row)-1:
                    cell.border = Border(
                            right = Side(
                                    border_style = gridStyle
                                )
                        )
            else:
                if  n == len(row)-1:
                    cell.border = Border(
                            bottom = Side(
                                    border_style = gridStyle
                                )
                        )
                else:
                    cell.border = Border(
                            bottom = Side(
                                    border_style = gridStyle
                                ),
                            right = Side(
                                    border_style = gridStyle
                                )
                        )

def generateRangeExpression(startRow = 0,
                            endRow = 0,
                            startColumn = "A",
                            endColumn = "A"):

    expression = "{startCol}{startRow}:{endCol}{endRow}".format(
            startCol = startColumn,
            startRow = startRow,
            endCol  = endColumn,
            endRow = endRow
        )
    return expression

def TamplateLayout(endRow):

    if endRow < 6:
        raise ValueError(
                "input of {num} is too small. The 'endRow' argument must be greater or equal than 6".format(
                    num = endRow
                )
            )
    elif type(endRow) != int:
        raise TypeError(
                "input of type {input_type_name} does not match the required type {type_name}".format(
                    input_type_name = type(endRow),
                    type_name = type(123)
                )
            )

    # create Workbook
    converter = ColorConverter()
    converter.interpretation = "openpyxl"
    converter.colorType = "HEX"
    wb = openpyxl.Workbook()

    # customize worksheet
    ws = wb.active
    ws.title = "Trainingsplan"
    ws.page_margins = openpyxl.worksheet.page.PageMargins(
            left = 0.25,
            right = 0.25,
            top = 0.75,
            bottom = 0.75,
            header = 0.5,
            footer = 0.5
        )
    ws.print_options = openpyxl.worksheet.page.PrintOptions(
            horizontalCentered = True,
            verticalCentered = False
        )

    # header
    ws.cell(2, 1, value = "Name").font = Font(
            b = True
        )
    # ws.cell(3, 1, value = "Julian")

    ws.cell(2, 4, value = "Trainingszeitraum").font = Font(
            b = True
        )
    ws.cell(3, 4, value = "Anfang:")
    # ws.cell(3, 6, value = "15.04.2020")
    ws.cell(4, 4, value = "Ende:")
    # ws.cell(4, 6, value = "15.04.2020")


    ws.cell(2, 8, value = "Trainingsmodus").font = Font(
            b = True
        )
    # ws.cell(3, 9, value = "Maximalkraft")

    ws.merge_cells("A2:B2")
    ws.merge_cells("A3:B3")
    ws.merge_cells("A6:B6")
    ws.merge_cells("D2:G2")
    ws.merge_cells("D3:E3")
    ws.merge_cells("D4:E4")
    ws.merge_cells("F3:G3")
    ws.merge_cells("F4:G4")
    ws.merge_cells("H2:J2")
    ws.merge_cells("H3:J3")
    setAlignment(ws, "A1:J5")
    setAlignment(ws, "D3:E3", horizontal = "left")
    setAlignment(ws, "D4:E4", horizontal = "left")
    setAlignment(ws, "F3:G3", horizontal = "right")
    setAlignment(ws, "F4:G4", horizontal = "right")
    setAlignment(ws, "H2:H2", horizontal = "right")
    setAlignment(ws, "H3:H3", horizontal = "right")
    setBorder(ws, "A1:J2")
    setBorder(ws, "A3:J5")
    setBorder(ws, "C2:C3", borderStyle = None)
    setBorder(ws, "H2:H3", borderStyle = None)
    setBackground(ws, "A1:J5",
                  fill_type = "solid",
                  fgColor = converter.color("white")
            )

    # table header
    ws.cell(6, 1, value = "Übung").font = Font(
            b = True
        )
    ws.cell(6, 3, value = "Sätze").font = Font(
            b = True
        )
    ws.cell(6, 4, value = "Whlg.").font = Font(
            b = True
        )
    ws.cell(6, 5, value = "W1").font = Font(
            b = True
        )
    ws.cell(6, 6, value = "W2").font = Font(
            b = True
        )
    ws.cell(6, 7, value = "W3").font = Font(
            b = True
        )
    ws.cell(6, 8, value = "W4").font = Font(
            b = True
        )
    ws.cell(6, 9, value = "W5").font = Font(
            b = True
        )
    ws.cell(6, 10, value = "W6").font = Font(
            b = True
        )
    setAlignment(ws, "A6:J6")
    setBorder(ws, "A6:J6")
    setBackground(ws, "A6:J6",
                  fill_type = "solid",
                  fgColor = converter.color("gray")
            )

    # table body
    startRow = 7
    endRow = endRow

    for i in range(startRow, endRow+1):
        cellRange = generateRangeExpression(
                startColumn = "A",
                endColumn = "B",
                startRow = i,
                endRow = i
            )
        ws.merge_cells(cellRange)

    setAlignment(ws, generateRangeExpression(
            startColumn = "A",
            endColumn = "A",
            startRow = startRow,
            endRow = endRow
        ),
            horizontal = "left",
            vertical = "center"
        )
    setAlignment(ws, generateRangeExpression(
            startColumn = "B",
            endColumn = "J",
            startRow = startRow,
            endRow = endRow
        ),
            horizontal = "center",
            vertical = "center"
        )
    setBorder(ws,generateRangeExpression(
            startColumn = "A",
            endColumn = "J",
            startRow = startRow,
            endRow = endRow
        ),
              borderStyle = "thick",
              gridStyle = "thin"
        )
    setBackground(ws, generateRangeExpression(
            startColumn = "A",
            endColumn = "D",
            startRow = startRow,
            endRow = endRow
        ),
                  fill_type = "solid",
                  fgColor = converter.color("gray")
        )
    return wb


if __name__ == "__main__":

    wb = TamplateLayout(40)
    ws = wb.active

    # save workbook in designated file
    wb.save("test_trainingroutine.xlsx")


