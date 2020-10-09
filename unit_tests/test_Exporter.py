# -*- coding: utf-8 -*-
"""
Created on Thu Apr 16 14:37:34 2020

@author: Julian
"""
import re
import unittest
import datetime
import openpyxl
import xlsxwriter
import tempfile
import pathlib
import numpy as np
from PyQt5.QtGui import QStandardItemModel

import MainModules.Database as db
from MainModules.Exporter import Exporter
from UtilityModules.CustomModel import CustomSqlModel
from UtilityModules.MiscUtilities import GetProjectRoot

class ExporterProperties(unittest.TestCase):

    def setUp(self):
        self.raiseTypeErrors = [
                123,
                123.123,
                [],
                (),
                {},
            ]
        self.raiseValueErrors = [
                "randomParent/randomPath/randomFile",
                ""
            ]
        self.projectRoot = GetProjectRoot()
        self.file = pathlib.Path("unit_tests/test_files/test_database_2.db")
        self.database = self.projectRoot / self.file
        self.currentDir = pathlib.Path().cwd()
        self.parentDir = pathlib.Path().cwd().parent
        self.exporter = Exporter()

    def test_instance(self):
        self.assertIsInstance(
                self.exporter, Exporter
            )

    def test_alternativeModel_getter(self):
        self.assertEqual(
                self.exporter.alternativeModel(), None
            )

    def test_database_getter(self):
        self.assertEqual(
                self.exporter.database(), None
            )

    def test_exportPath_getter(self):
        self.assertEqual(
                self.exporter.exportPath(), None
            )

    def test_databaseName_getter(self):
        self.assertEqual(
                self.exporter.databaseName(), None
            )

    def test_databasePath_getter(self):
        self.assertEqual(
                self.exporter.databasePath(), None
            )

    def test_layoutProperties_getter(self):
        props = {
                "headerStartRow": 0,
                "routineStartRow": int(),
                "alternativeStartRow": int(),
                "layoutMaxRows": 40,
                "layoutMaxCols": 10
            }

        self.assertEqual(
                self.exporter.layoutProperties(), props
            )

    def test_name_getter(self):
        self.assertEqual(
                self.exporter.name(), None
            )

    def test_noteModel_getter(self):
        self.assertEqual(
                self.exporter.noteModel(), None
            )

    def test_routineName_getter(self):
        self.assertEqual(
                self.exporter.routineName(), None
            )

    def test_routineModel_getter(self):
        self.assertEqual(
                self.exporter.routineModel(), None
            )

    def test_trainingMode_getter(self):
        self.assertEqual(
                self.exporter.trainingMode(), None
            )

    def test_trainingPeriode_getter(self):
        self.assertEqual(
                self.exporter.trainingPeriode(), list()
            )

    def test_workBook_getter(self):
        self.assertEqual(
                self.exporter.workBook(), None
            )

    def test_workSheet_getter(self):
        self.assertEqual(
                self.exporter.workSheet(), None
            )

    def test_alternativeModel_setter(self):
        model = CustomSqlModel()
        self.exporter.setAlternativeModel(model)

        self.assertEqual(
                self.exporter.alternativeModel(), model
            )

        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setModel, val
                    )

        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setModel, val
                    )


    def test_name_setter(self):
        self.exporter.setName("Test Name")
        self.assertEqual(self.exporter.name(), "Test Name")
        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setName, val
                    )
        self.assertRaises(
                ValueError, self.exporter.setName, self.raiseValueErrors[1]
            )

    def test_database_setter(self):
        self.exporter.setDatabase(self.database)
        self.assertEqual(
                self.exporter.database(), self.database
            )
        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(TypeError, self.exporter.setDatabase, val)
        for val in self.raiseValueErrors:
            with self.subTest(val = val):
                self.assertRaises(ValueError, self.exporter.setDatabase, val)

    def test_exportPath_setter(self):
        self.exporter.setExportPath(self.currentDir)
        self.assertEqual(
                self.exporter.exportPath(), self.currentDir
            )
        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setExportPath, val
                    )
        for val in self.raiseValueErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        ValueError, self.exporter.setExportPath, val
                    )

    def test_layoutProperties_setter(self):
        props = {
                "headerStartRow": 0,
                "routineStartRow": 1,
                "alternativeStartRow": 2,
                "layoutMaxRows": 3,
                "layoutMaxCols": 4
            }

        raiseTypeErrors = self.raiseTypeErrors
        del raiseTypeErrors[4]
        raiseTypeErrors.append("")

        self.exporter.setLayoutProperties(props)
        self.assertEqual(
                self.exporter.layoutProperties(), props
            )
        for val in raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setLayoutProperties, val
                    )

    def test_noteModel_setter(self):
        model = QStandardItemModel()
        self.exporter.setNoteModel(model)

        self.assertEqual(
                self.exporter.noteModel(), model
            )

    def test_routineName_setter(self):
        self.exporter.setRoutineName("Test_Name")
        self.assertEqual(
                self.exporter.routineName(), "Test_Name.xlsx"
            )
        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setRoutineName, val
                    )
        self.assertRaises(
                ValueError, self.exporter.setRoutineName, self.raiseValueErrors[1]
            )

    def test_routineModel_setter(self):
        model = CustomSqlModel()
        self.exporter.setRoutineModel(model)

        self.assertEqual(
                self.exporter.routineModel(), model
            )

    def test_trainingMode_setter(self):
        raiseTypeErrors = self.raiseTypeErrors
        raiseValueErrors = self.raiseValueErrors[1]

        self.exporter.setTrainingMode("TestMode")
        self.assertEqual(
                self.exporter.trainingMode(), "TestMode"
            )

        for val in raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setTrainingMode, val
                    )

        for val in raiseValueErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        ValueError, self.exporter.setTrainingMode, val
                    )

    def test_trainingPeriode_setter(self):
        startDateValues = [
                datetime.date(year, month, day)
                for year in range(2019,2021)
                for month in range(1,3)
                for day in range(15,17)
            ]
        endDateValues = [
                start + datetime.timedelta(days = 42)
                for start in startDateValues
            ]
        exceptionValues = ["", 123.124, (), [], {}]

        for i, startDate in enumerate(startDateValues):
            with self.subTest(startDate = startDate):
                endDate = endDateValues[i]
                self.exporter.setTrainingPeriode(
                        startDate.year, startDate.month, startDate.day
                    )
                self.assertEqual(
                        self.exporter.trainingPeriode(), [
                                startDate.strftime("%d.%m.%Y"),
                                endDate.strftime("%d.%m.%Y")
                            ]
                    )
        for value in exceptionValues:
            with self.subTest(value = value):
                self.assertRaises(
                        TypeError, self.exporter.setTrainingPeriode, value, 2, 3
                    )
        for value in exceptionValues:
            with self.subTest(value = value):
                self.assertRaises(
                        TypeError, self.exporter.setTrainingPeriode, 1, value, 3
                    )
        for value in exceptionValues:
            with self.subTest(value = value):
                self.assertRaises(
                        TypeError, self.exporter.setTrainingPeriode, 1, 2, value
                    )

    def test_workBook_setter(self):
        raiseTypeErrors = self.raiseTypeErrors
        raiseTypeErrors.append("")
        wb = xlsxwriter.Workbook()

        self.exporter.setWorkBook(wb)
        self.assertEqual(
                self.exporter.workBook(), wb
            )
        for val in raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setWorkBook, val
                    )

    def test_workSheet_setter(self):
        wb = xlsxwriter.Workbook()
        ws = wb.add_worksheet()

        self.exporter.setWorkSheet(ws)
        self.assertEqual(
                self.exporter.workSheet(), ws
            )

        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.workSheet(), val
                    )

    def tearDown(self):
        pass

class TrainingRoutineLayout(unittest.TestCase):

    def setUp(self):
        self.raiseTypeErrors = [
                123,
                123.123,
                [],
                (),
                {},
            ]
        self.raiseValueErrors = [
                "randomParent/randomPath/randomFile",
                ""
            ]
        self.projectRoot = GetProjectRoot()
        self.database = self.projectRoot / pathlib.Path("unit_tests/test_files/test_database_2.db")
        self.exportPath = self.projectRoot / pathlib.Path("unit_tests/test_files")
        self.databaseName = self.database.stem
        self.routineName = "test_database_2.xlsx"
        self.name = "Aphrodite"
        self.trainingMode = "TestMode"
        self.columnCount = 10
        self.rowCountValues = [40, 60, 80, 100]

        self.db = db.database(self.database)

        self.alternativeModel = CustomSqlModel(
                database = self.db,
                table = "training_alternatives",
                valueStartIndex = 0,
                tableStartIndex = 0
                )
        self.alternativeModel.populateModel()

        self.noteModel = CustomSqlModel(
                database = self.db,
                table = "training_notes",
                valueStartIndex = 0,
                tableStartIndex = 0
                )
        self.noteModel.populateModel()

        self.routineModel = CustomSqlModel(
                database = self.db,
                table = "training_routine",
                valueStartIndex = 0,
                tableStartIndex = 0
                )
        self.routineModel.populateModel()

        self.exporter = Exporter()
        self.exporter.setDatabase(self.database)
        self.exporter.setAlternativeModel(self.alternativeModel)
        self.exporter.setNoteModel(self.noteModel)
        self.exporter.setRoutineModel(self.routineModel)
        self.exporter.setName(self.name)
        self.exporter.setRoutineName(self.routineName)
        self.exporter.setTrainingMode(self.trainingMode)
        self.exporter.setTrainingPeriode(2020, 11, 11)
        self.exporter.setExportPath(self.exportPath)
        self.exporter.export()

    def test_workbook_validity(self):
        wb = self.exporter.workBook()
        self.assertIsInstance(wb, xlsxwriter.Workbook)

    def test_dataFromDatabase(self):
        routineData, alternativeData, noteData = self.exporter.dataFromDatabase(self.database)

        # test routine data
        self.assertEqual(
                len(np.array(routineData).shape), 2
            )

        self.assertEqual(
                len(routineData[0]), 11
            )

        # test alternative data
        self.assertEqual(
                len(np.array(alternativeData).shape), 2
            )

        self.assertEqual(
                len(alternativeData[0]), 14
            )

        # test note data
        self.assertEqual(
                len(np.array(noteData).shape), 2
            )

        self.assertEqual(
                len(noteData[0]), 4
            )

    def test_dataFromModel(self):
        routineData, alternativeData, noteData = self.exporter.dataFromModel()

        # test routine data
        self.assertEqual(
                len(np.array(routineData).shape), 2
            )

        self.assertEqual(
                len(routineData[0]), 11
            )

        # test alternative data
        self.assertEqual(
                len(np.array(alternativeData).shape), 2
            )

        self.assertEqual(
                len(alternativeData[0]), 14
            )

        # test note data
        self.assertEqual(
                len(np.array(noteData).shape), 2
            )

        self.assertEqual(
                len(noteData[0]), 4
            )

    def test_layout_dimensions(self):
        oldPath = self.exporter.exportPath()
        oldName = self.exporter.routineName()
        data = self.db.data("training_routine")

        with tempfile.TemporaryDirectory() as tmpdirname:
            self.exporter.setRoutineName("test_layout_dimensions")
            self.exporter.setExportPath(tmpdirname)
            for row in self.rowCountValues:
                testProps = {
                        "headerStartRow": 0,
                        "routineStartRow": 7,
                        "alternativeStartRow": 7 + len(data) + 3, # 7: routine Start row, len(data): length of training_routine, 3: space between training_routine and training_alternatives
                        "layoutMaxRows": row,
                        "layoutMaxCols": 10
                    }
                props = self.exporter.layoutProperties()
                props["layoutMaxRows"] = row
                self.exporter.setLayoutProperties(props)
                self.exporter.export()

                props = self.exporter.layoutProperties()
                self.assertEqual(
                        props, testProps
                    )

        self.exporter.setExportPath(oldPath)
        self.exporter.setRoutineName(oldName)

    def test_layout_elements(self):
        wb = openpyxl.load_workbook(
                filename = self.exporter.exportPath() / self.exporter.routineName()
            )
        ws = wb.active

        cell = ws["A2"]
        self.assertEqual(cell.value, "Name")
        self.assertEqual(cell.border.bottom.style, "medium")
        self.assertTrue(cell.font.b)

        cell = ws["D2"]
        self.assertEqual(cell.value, "Trainingszeitraum")
        self.assertEqual(cell.border.bottom.style, "medium")
        self.assertTrue(cell.font.b)

        cell = ws["D3"]
        self.assertEqual(cell.value, "Anfang:")
        self.assertTrue(cell.font.b)

        cell = ws["D4"]
        self.assertEqual(cell.value, "Ende:")
        self.assertTrue(cell.font.b)

        cell = ws["I2"]
        self.assertEqual(cell.value, "Trainingsmodus")
        self.assertTrue(cell.font.b)
        self.assertEqual(cell.border.bottom.style, "medium")

        cell = ws["A7"]
        self.assertEqual(cell.value, "Übung")
        self.assertTrue(cell.font.b)
        cell = ws["C7"]
        self.assertEqual(cell.value, "Sätze")
        self.assertTrue(cell.font.b)
        cell = ws["D7"]
        self.assertEqual(cell.value, "Whlg.")
        self.assertTrue(cell.font.b)
        cell = ws["E7"]
        self.assertEqual(cell.value, "W1")
        self.assertTrue(cell.font.b)
        cell = ws["F7"]
        self.assertEqual(cell.value, "W2")
        self.assertTrue(cell.font.b)
        cell = ws["G7"]
        self.assertEqual(cell.value, "W3")
        self.assertTrue(cell.font.b)
        cell = ws["H7"]
        self.assertEqual(cell.value, "W4")
        self.assertTrue(cell.font.b)
        cell = ws["I7"]
        self.assertEqual(cell.value, "W5")
        self.assertTrue(cell.font.b)
        cell = ws["J7"]
        self.assertEqual(cell.value, "W6")
        self.assertTrue(cell.font.b)
        for i in range(1, self.columnCount + 1):
            with self.subTest(i = i):
                cell = ws.cell(7, i)
                if i == 1:
                    self.assertEqual(cell.border.left.style, "medium")
                    self.assertEqual(cell.border.top.style, "medium")
                    self.assertEqual(cell.border.bottom.style, "medium")
                elif i == 2: # skip the merged part of the cell
                    continue
                elif i == self.columnCount:
                    self.assertEqual(cell.border.right.style, "medium")
                    self.assertEqual(cell.border.top.style, "medium")
                    self.assertEqual(cell.border.bottom.style, "medium")
                else:
                    self.assertEqual(cell.border.top.style, "medium")
                    self.assertEqual(cell.border.bottom.style, "medium")
                    self.assertEqual(cell.fill.fgColor.rgb, "FF808080")

    def test_table_body(self):
        grayCols = [1, 3, 4]
        witheCols = [5, 6, 7, 8, 9, 10]
        oldPath = self.exporter.exportPath()
        oldName = self.exporter.routineName()

        def do_the_magic():
            for rowCount in self.rowCountValues:
                props = self.exporter.layoutProperties()
                props["layoutMaxRows"] = rowCount
                self.exporter.setLayoutProperties(props)
                self.exporter.export()
                wb = openpyxl.load_workbook(
                        filename = self.exporter.exportPath() / self.exporter.routineName()
                    )
                ws = wb.active

                with self.subTest(rowCount = rowCount):
                    for row in range(8, rowCount + 1):
                        with self.subTest(row = row):
                            for col in grayCols:
                                with self.subTest(col = col):
                                    cell = ws.cell(row, col)
                                    self.assertEqual(cell.fill.fgColor.rgb, "FF808080")
                            for col in witheCols:
                                with self.subTest(col = col):
                                    cell = ws.cell(row, col)
                                    self.assertEqual(cell.fill.fgColor.rgb, "FFFFFFFF") # FFFFFFFF

                    for row in range(8, rowCount + 1):
                        with self.subTest(row = row):
                            for col in range(1, self.columnCount + 1):
                                with self.subTest(col = col):
                                    cell = ws.cell(row, col)
                                    if row != rowCount:
                                        if col == 1:
                                            self.assertEqual(cell.border.left.style, "medium")
                                            self.assertEqual(cell.border.bottom.style, "thin")
                                            self.assertEqual(cell.border.right.style, "thin")
                                        elif col == self.columnCount:
                                            self.assertEqual(cell.border.right.style, "medium")
                                            self.assertEqual(cell.border.bottom.style, "thin")
                                        else:
                                            self.assertEqual(cell.border.right.style, "thin")
                                            self.assertEqual(cell.border.bottom.style, "thin")
                                    else:
                                        if col == 1:
                                            self.assertEqual(cell.border.left.style, "medium")
                                            self.assertEqual(cell.border.bottom.style, "medium")
                                            self.assertEqual(cell.border.right.style, "thin")
                                        elif col == self.columnCount:
                                            self.assertEqual(cell.border.right.style, "medium")
                                            self.assertEqual(cell.border.bottom.style, "medium")
                                        else:
                                            self.assertEqual(cell.border.bottom.style, "medium")
                                            self.assertEqual(cell.border.right.style, "thin")

        with tempfile.TemporaryDirectory() as tmpdirname:
            self.exporter.setExportPath(tmpdirname)
            self.exporter.setRoutineName("test_table_body")
            do_the_magic()

        self.exporter.setExportPath(oldPath)
        self.exporter.setRoutineName(oldName)

    def test_populateRoutine_from_database(self):
        oldPath = self.exporter.exportPath()
        oldName = self.exporter.routineName()

        def do_the_magic():
            routineData,_,_ = self.exporter.dataFromDatabase()
            data = self.db.data("training_routine", self.databaseName)
            wb = openpyxl.load_workbook(
                    filename = self.exporter.exportPath() / self.exporter.routineName()
                )
            ws = wb.active
            headerTestCells = ["A3", "F3", "F4", "H3"]
            headerTestResults = [
                            self.name,
                            self.exporter.trainingPeriode()[0],
                            self.exporter.trainingPeriode()[1],
                            self.trainingMode
                         ]
            for i, val in enumerate(headerTestCells):
                with self.subTest(val = val):
                    self.assertEqual(
                            ws[val].value, headerTestResults[i]
                        )

            testValues = [data[i][0] for i in range(len(data))]
            for i, val in enumerate(testValues):
                with self.subTest(i = i):
                    cell = ws["A" + str(8 + i)]
                    value = cell.value[:len(val)]
                    self.assertEqual(
                            value, val
                        )

            testValues = [int(data[i][1]) for i in range(len(data))]
            for i, val in enumerate(testValues):
                with self.subTest(i = i):
                    cell = ws["C" + str(8 + i)]
                    self.assertEqual(
                            cell.value, val
                        )

            testValues = [int(data[i][2]) for i in range(len(data))]
            for i, val in enumerate(testValues):
                with self.subTest(i = i):
                    cell = ws["D" + str(8 + i)]
                    self.assertEqual(
                            cell.value, val
                        )

        with tempfile.TemporaryDirectory() as tmpdirname:
            self.exporter.setExportPath(tmpdirname)
            self.exporter.setRoutineName("test_populateRoutine_from_database.xlsx")
            self.exporter.export()
            do_the_magic()

        self.exporter.setExportPath(oldPath)
        self.exporter.setRoutineName(oldName)

    def test_populateRoutine_from_model(self):
        oldPath = self.exporter.exportPath()
        oldName = self.exporter.routineName()

        def do_the_magic():
            routineData,_,_ = self.exporter.dataFromModel()
            data = self.db.data("training_routine", self.databaseName)
            wb = openpyxl.load_workbook(
                    filename = self.exporter.exportPath() / self.exporter.routineName()
                )
            ws = wb.active

            headerTestCells = ["A3", "F3", "F4", "H3"]
            headerTestResults = [
                            self.name,
                            self.exporter.trainingPeriode()[0],
                            self.exporter.trainingPeriode()[1],
                            self.trainingMode
                         ]
            for i, val in enumerate(headerTestCells):
                with self.subTest(val = val):
                    self.assertEqual(
                            ws[val].value, headerTestResults[i]
                        )

            testValues = [data[i][0] for i in range(len(data))]
            for i, val in enumerate(testValues):
                with self.subTest(i = i):
                    cell = ws["A" + str(8 + i)]
                    value = cell.value[:len(val)]
                    self.assertEqual(
                            value, val
                        )

            testValues = [int(data[i][1]) for i in range(len(data))]
            for i, val in enumerate(testValues):
                with self.subTest(i = i):
                    cell = ws["C" + str(8 + i)]
                    self.assertEqual(
                            cell.value, val
                        )

            testValues = [int(data[i][2]) for i in range(len(data))]
            for i, val in enumerate(testValues):
                with self.subTest(i = i):
                    cell = ws["D" + str(8 + i)]
                    self.assertEqual(
                            cell.value, val
                        )

        with tempfile.TemporaryDirectory() as tmpdirname:
            self.exporter.setExportPath(tmpdirname)
            self.exporter.setRoutineName("test_populateRoutine_from_database.xlsx")
            self.exporter.export()
            do_the_magic()

        self.exporter.setExportPath(oldPath)
        self.exporter.setRoutineName(oldName)

    def tearDown(self):
        exportFilePath = self.exportPath / pathlib.Path(self.routineName)

        try:
            exportFilePath.unlink()
        except FileNotFoundError:
            pass

if __name__ == "__main__":
    unittest.main()
