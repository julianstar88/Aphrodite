# -*- coding: utf-8 -*-
"""
Created on Thu Apr 16 14:37:34 2020

@author: Julian
"""

import unittest
import examples.SQLite3_Database.Database as db
import datetime
import openpyxl
import os
import pathlib2
from MainModules.Exporter import Exporter

class ExporterProperties(unittest.TestCase):

    def setUp(self):
        self.raiseTypeErrors = [
                123,
                123.123,
                [],
                (),
                {},
            ]
        self.raiseValueErrors = [
                "randomParent/randomPath/randomFile",
                ""
            ]
        pathObj = pathlib2.Path().cwd() / "temp_test_database.db"
        self.path = str(pathObj)
        self.db = db.database(pathObj.parent)
        self.db.createDatabase("temp_test_database")
        self.exporter = Exporter()

    def test_instance(self):
        self.assertIsInstance(
                self.exporter, Exporter
            )

    def test_database_getter(self):
        self.assertEqual(
                self.exporter.database(), None
            )

    def test_exportPath_getter(self):
        self.assertEqual(
                self.exporter.exportPath(), None
            )

    def test_name_getter(self):
        self.assertEqual(
                self.exporter.name(), None
            )

    def test_routineName_getter(self):
        self.assertEqual(
                self.exporter.routineName(), None
            )

    def test_trainingPeriode_getter(self):
        self.assertEqual(
                self.exporter.trainingPeriode(), [None, None]
            )

    def test_name_setter(self):
        self.exporter.setName("Test Name")
        self.assertEqual(self.exporter.name(), "Test Name")
        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setName, val
                    )
        self.assertRaises(
                ValueError, self.exporter.setName, self.raiseValueErrors[1]
            )

    def test_database_setter(self):
        self.exporter.setDatabase(self.path)
        self.assertEqual(
                self.exporter.database(), self.path
            )
        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(TypeError, self.exporter.setDatabase, val)
        for val in self.raiseValueErrors:
            with self.subTest(val = val):
                self.assertRaises(ValueError, self.exporter.setDatabase, val)

    def test_exportPath_setter(self):
        pathObj = pathlib2.Path().cwd()
        self.exporter.setExportPath(str(pathObj))
        self.assertEqual(
                self.exporter.exportPath(), str(pathObj)
            )
        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setExportPath, val
                    )
        for val in self.raiseValueErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        ValueError, self.exporter.setExportPath, val
                    )

    def test_routineName_setter(self):
        self.exporter.setRoutineName("Test Name")
        self.assertEqual(
                self.exporter.routineName(), "Test Name"
            )
        for val in self.raiseTypeErrors:
            with self.subTest(val = val):
                self.assertRaises(
                        TypeError, self.exporter.setRoutineName, val
                    )
        self.assertRaises(
                ValueError, self.exporter.setRoutineName, self.raiseValueErrors[1]
            )

    def test_trainingPeriode_setter(self):
        startDateValues = [
                datetime.date(year, month, day)
                for year in range(2019,2021)
                for month in range(1,3)
                for day in range(15,17)
            ]
        endDateValues = [
                start + datetime.timedelta(days = 42)
                for start in startDateValues
            ]
        exceptionValues = ["", 123.124, (), [], {}]

        for i, startDate in enumerate(startDateValues):
            with self.subTest(startDate = startDate):
                endDate = endDateValues[i]
                self.exporter.setTrainingPeriode(
                        startDate.year, startDate.month, startDate.day
                    )
                self.assertEqual(
                        self.exporter.trainingPeriode(), (startDate, endDate)
                    )
        for value in exceptionValues:
            with self.subTest(value = value):
                self.assertRaises(
                        TypeError, self.exporter.setTrainingPeriode, value, 2, 3
                    )
        for value in exceptionValues:
            with self.subTest(value = value):
                self.assertRaises(
                        TypeError, self.exporter.setTrainingPeriode, 1, value, 3
                    )
        for value in exceptionValues:
            with self.subTest(value = value):
                self.assertRaises(
                        TypeError, self.exporter.setTrainingPeriode, 1, 2, value
                    )

    def tearDown(self):
        os.remove(self.path)

class TrainingRoutineLayout(unittest.TestCase):

    def setUp(self):
        self.columnCount = 10
        self.rowCountValues = [6, 10, 20, 40, 60]
        self.exporter = Exporter()

    def test_workbook_validity(self):
        wb = self.exporter.routineLayout(self.rowCountValues[1])
        self.assertIsInstance(wb, openpyxl.Workbook)

    def test_layout_dimenstions(self):
        for rowCount in self.rowCountValues:

            wb = self.exporter.routineLayout(rowCount)
            ws = wb.active

            with self.subTest(rowCount = rowCount):
                for i in range(1, self.columnCount+1):
                    with self.subTest(i = i):
                        cell = ws.cell(1, i)
                        self.assertEqual(cell.border.top.style, "thick")

                for i in range(1, rowCount+1):
                    with self.subTest(i = i):
                        cell = ws.cell(i,10)
                        self.assertEqual(cell.border.right.style, "thick")

    def test_layout_elements(self):
        wb = self.exporter.routineLayout(self.rowCountValues[1])
        ws = wb.active

        cell = ws["A2"]
        self.assertEqual(cell.value, "Name")
        self.assertEqual(cell.border.bottom.style, "thick")
        self.assertTrue(cell.font.b)

        cell = ws["D2"]
        self.assertEqual(cell.value, "Trainingszeitraum")
        self.assertEqual(cell.border.bottom.style, "thick")
        self.assertTrue(cell.font.b)

        cell = ws["D3"]
        self.assertEqual(cell.value, "Anfang:")
        self.assertFalse(cell.font.b)

        cell = ws["D4"]
        self.assertEqual(cell.value, "Ende:")
        self.assertFalse(cell.font.b)

        cell = ws["I2"]
        self.assertEqual(cell.value, "Trainingsmodus")
        self.assertEqual(cell.border.bottom.style, "thick")
        self.assertTrue(cell.font.b)

        cell = ws["A6"]
        self.assertEqual(cell.value, "Übung")
        self.assertTrue(cell.font.b)
        cell = ws["C6"]
        self.assertEqual(cell.value, "Sätze")
        self.assertTrue(cell.font.b)
        cell = ws["D6"]
        self.assertEqual(cell.value, "Whlg.")
        self.assertTrue(cell.font.b)
        cell = ws["E6"]
        self.assertEqual(cell.value, "W1")
        self.assertTrue(cell.font.b)
        cell = ws["F6"]
        self.assertEqual(cell.value, "W2")
        self.assertTrue(cell.font.b)
        cell = ws["G6"]
        self.assertEqual(cell.value, "W3")
        self.assertTrue(cell.font.b)
        cell = ws["H6"]
        self.assertEqual(cell.value, "W4")
        self.assertTrue(cell.font.b)
        cell = ws["I6"]
        self.assertEqual(cell.value, "W5")
        self.assertTrue(cell.font.b)
        cell = ws["J6"]
        self.assertEqual(cell.value, "W6")
        self.assertTrue(cell.font.b)
        for i in range(1, self.columnCount+1):
            with self.subTest(i = i):
                cell = ws.cell(6, i)
                if i == 1:
                    self.assertEqual(cell.border.left.style, "thick")
                    self.assertEqual(cell.border.top.style, "thick")
                    self.assertEqual(cell.border.bottom.style, "thick")
                elif i == self.columnCount:
                    self.assertEqual(cell.border.right.style, "thick")
                    self.assertEqual(cell.border.top.style, "thick")
                    self.assertEqual(cell.border.bottom.style, "thick")
                self.assertEqual(cell.border.top.style, "thick")
                self.assertEqual(cell.border.bottom.style, "thick")
                self.assertEqual(cell.fill.fgColor.rgb, "00808080")

    def test_table_body(self):
        grayCols = [1, 2, 3, 4]
        witheCols = [5, 6, 7, 8, 9, 10]

        for rowCount in self.rowCountValues:
            wb = self.exporter.routineLayout(rowCount)
            ws = wb.active

            with self.subTest(rowCount = rowCount):
                for row in range(7, rowCount+1):
                    with self.subTest(row = row):
                        for col in grayCols:
                            with self.subTest(col = col):
                                cell = ws.cell(row, col)
                                self.assertEqual(cell.fill.fgColor.rgb, "00808080")
                        for col in witheCols:
                            with self.subTest(col = col):
                                cell = ws.cell(row, col)
                                self.assertEqual(cell.fill.fgColor.rgb, "00000000")

                for row in range(7, rowCount+1):
                    with self.subTest(row = row):
                        for col in range(1, self.columnCount+1):
                            with self.subTest(col = col):
                                cell = ws.cell(row, col)
                                if row != rowCount:
                                    if col == 1:
                                        self.assertEqual(cell.border.left.style, "thick")
                                        self.assertEqual(cell.border.bottom.style, "thin")
                                        self.assertEqual(cell.border.right.style, "thin")
                                    elif col == self.columnCount:
                                        self.assertEqual(cell.border.right.style, "thick")
                                        self.assertEqual(cell.border.bottom.style, "thin")
                                    else:
                                        self.assertEqual(cell.border.right.style, "thin")
                                        self.assertEqual(cell.border.bottom.style, "thin")
                                else:
                                    if col == 1:
                                        self.assertEqual(cell.border.left.style, "thick")
                                        self.assertEqual(cell.border.bottom.style, "thick")
                                        self.assertEqual(cell.border.right.style, "thin")
                                    elif col == self.columnCount:
                                        self.assertEqual(cell.border.right.style, "thick")
                                        self.assertEqual(cell.border.bottom.style, "thick")
                                    else:
                                        self.assertEqual(cell.border.bottom.style, "thick")
                                        self.assertEqual(cell.border.right.style, "thin")

    def test_routine_population(self):
        wb = self.exporter.routineLayout()
        self.exporter.setName("Aphrodite")
        self.exporter.populateRoutine(wb)
        ws = wb.active



if __name__ == "__main__":
    unittest.main()
